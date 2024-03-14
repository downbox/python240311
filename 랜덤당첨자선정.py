import sys
import mysql.connector
import pandas as pd
from PyQt5.QtWidgets import QApplication, QWidget, QDialog, QPushButton, QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem, QHeaderView, QLineEdit, QLabel, QDateTimeEdit, QMessageBox
from PyQt5.QtCore import pyqtSlot, QDateTime
from openpyxl import load_workbook
import pyperclip
import os

class LoginDialog(QDialog):
    def __init__(self, database_config):
        super().__init__()
        self.database_config = database_config
        self.setupUI()

    def setupUI(self):
        self.setWindowTitle("로그인")
        layout = QVBoxLayout(self)

        # ID 입력
        self.idLabel = QLabel("ID:")
        self.idInput = QLineEdit()
        layout.addWidget(self.idLabel)
        layout.addWidget(self.idInput)

        # Password 입력
        self.pwLabel = QLabel("Password:")
        self.pwInput = QLineEdit()
        self.pwInput.setEchoMode(QLineEdit.Password)
        layout.addWidget(self.pwLabel)
        layout.addWidget(self.pwInput)

        # 로그인 버튼
        self.loginButton = QPushButton("로그인")
        self.loginButton.clicked.connect(self.check_credentials)
        layout.addWidget(self.loginButton)

    def check_credentials(self):
        id_ = self.idInput.text()
        pw_ = self.pwInput.text()

        # 데이터베이스 접속
        try:
            conn = mysql.connector.connect(**self.database_config)
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM aptspace.user WHERE id = %s AND password = %s", (id_, pw_))
            result = cursor.fetchone()
            cursor.close()
            conn.close()

            if result:
                self.accept()  # 로그인 성공
            else:
                QMessageBox.warning(self, "오류", "ID나 Password가 잘못되었습니다.")
        except mysql.connector.Error as err:
            QMessageBox.warning(self, "오류", f"데이터베이스 연결 실패: {err}")

class LotteryApp(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()
        
    def initUI(self):
        self.setWindowTitle('기흥효성해링턴 랜덤추첨 프로그램')
        self.layout = QVBoxLayout()
        self.setLayout(self.layout)
        
        # 상단 버튼 툴바 레이아웃
        self.toolbarLayout = QHBoxLayout()
        
        # '추첨시작' 버튼
        self.btnDraw = QPushButton('추첨시작', self)
        self.btnDraw.clicked.connect(self.draw)
        self.toolbarLayout.addWidget(self.btnDraw)
        
        # '복사하기' 버튼
        self.btnCopy = QPushButton('복사하기', self)
        self.btnCopy.clicked.connect(self.copyResults)
        self.toolbarLayout.addWidget(self.btnCopy)
        
        # '초기화' 버튼
        self.btnReset = QPushButton('초기화', self)
        self.btnReset.clicked.connect(self.resetFields)
        self.toolbarLayout.addWidget(self.btnReset)
        
        # 입력 필드들을 담을 레이아웃
        self.inputsLayout = QHBoxLayout()
        
        # '추첨자수' 입력 필드
        self.numWinnersLabel = QLabel('추첨자수:')
        self.inputsLayout.addWidget(self.numWinnersLabel)
        self.numWinnersInput = QLineEdit(self)
        self.inputsLayout.addWidget(self.numWinnersInput)
        
        # '추첨자' 입력 필드
        self.entrantLabel = QLabel('추첨자:')
        self.inputsLayout.addWidget(self.entrantLabel)
        self.entrantInput = QLineEdit(self)
        self.inputsLayout.addWidget(self.entrantInput)

        # '추첨내용' 입력 필드 추가
        self.drawContentLabel = QLabel('추첨내용:')
        self.inputsLayout.addWidget(self.drawContentLabel)
        self.drawContentInput = QLineEdit(self)
        self.inputsLayout.addWidget(self.drawContentInput)

        # '추첨일시' 출력 필드
        self.drawDateLabel = QLabel('추첨일시:')
        self.inputsLayout.addWidget(self.drawDateLabel)
        self.drawDateTime = QDateTimeEdit(self)
        self.drawDateTime.setDisplayFormat("yyyy-MM-dd HH:mm:ss")
        self.drawDateTime.setReadOnly(True)
        self.inputsLayout.addWidget(self.drawDateTime)
        
        self.layout.addLayout(self.toolbarLayout)
        self.layout.addLayout(self.inputsLayout)
        
        self.table = QTableWidget()
        self.layout.addWidget(self.table)

        self.setGeometry(100, 100, 800, 400)  # 화면 가로 크기 확장
        self.show()

    @pyqtSlot()
    def draw(self):
        num_selections = int(self.numWinnersInput.text())
        entrant = self.entrantInput.text()
        draw_content = self.drawContentInput.text().strip()
        current_time = QDateTime.currentDateTime()
        self.drawDateTime.setDateTime(current_time)

        # '추첨내용'과 '추첨자' 필드가 비어 있는지 검증
        if not draw_content or len(draw_content) > 20 or not entrant:
            QMessageBox.warning(self, "입력 오류", "추첨내용과 추첨자를 입력하세요. 추첨내용은 20자 이내여야 합니다.")
            return
        
        # TIMESTAMP_추첨내용_RESULT 형식으로 파일명 생성
        timestamp = current_time.toString("yyyyMMddHHmmss")
        filename_suffix = draw_content if len(draw_content) <= 20 else draw_content[:20]
        excel_filename = f"{timestamp}_{filename_suffix}_RESULT.xlsx"

        try:
            # 엑셀 파일(.xlsx)에서 데이터를 읽습니다.
            data = pd.read_excel('C:/random_result/list.xlsx', engine='openpyxl')
            if num_selections > len(data):
                raise ValueError("선택한 숫자가 데이터 범위를 초과하였습니다.")
            
            self.result = data.sample(n=num_selections)
            self.showResults(self.result)

            # 엑셀에 추첨 결과 기록
            excel_path = os.path.join('C:/random_result', excel_filename)
            if os.path.exists(excel_path):
                book = load_workbook(excel_path)
            else:
                book = None

            with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a' if book else 'w') as writer:
                if book:
                    writer.book = book
                    writer.sheets = {ws.title: ws for ws in book.worksheets}
                
                self.result.to_excel(writer, index=False, sheet_name='Results', startrow=writer.sheets['Results'].max_row if 'Results' in writer.sheets else 0)
                # 추첨자수, 추첨자, 추첨일시 하단에 기록
                df_footer = pd.DataFrame({
                    '추첨자수': [num_selections],
                    '추첨자': [entrant],
                    '추첨일시': [current_time.toString("yyyy-MM-dd HH:mm:ss")]
                })
                df_footer.to_excel(writer, index=False, sheet_name='Footer', startrow=writer.book['Footer'].max_row if 'Footer' in writer.sheets else 0)
        except Exception as e:
            print(f"추첨 오류: {e}")

    def showResults(self, data):
        self.table.setRowCount(len(data))
        self.table.setColumnCount(len(data.columns))
        self.table.setHorizontalHeaderLabels(data.columns)

        for i, row in data.iterrows():
            for j, value in enumerate(row):
                self.table.setItem(i, j, QTableWidgetItem(str(value)))

        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

    def copyResults(self):
        selection = self.table.selectedIndexes()
        if selection:
            rows = sorted(index.row() for index in selection)
            if rows:
                clipboard = ""
                for row in rows:
                    row_data = []
                    for col in range(self.table.columnCount()):
                        item = self.table.item(row, col)
                        row_data.append(item.text())
                    clipboard += "\t".join(row_data) + "\n"
                pyperclip.copy(clipboard)

    
    def resetFields(self):
        # 필드 초기화
        self.numWinnersInput.clear()
        self.entrantInput.clear()
        self.drawDateTime.clear()
        # 결과 화면 초기화
        self.table.clearContents()
        self.table.setRowCount(0)

if __name__ == '__main__':
    app = QApplication(sys.argv)

    # 데이터베이스 설정
    db_config = {
        'user': 'aptspace',
        'password': 'dnaustkd12#',
        'host': 'my8003.gabiadb.com',
        'port': 3306,
        'database': 'aptspace',
        'raise_on_warnings': True,
    }

    # 로그인 대화상자 표시
    login_dialog = LoginDialog(db_config)
    if login_dialog.exec_() == QDialog.Accepted:
        main_window = LotteryApp()
        main_window.show()
    sys.exit(app.exec_())
